import pickle
from fizz_buzz import Config, Util
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

train_model_path = './src/model'
results_path = './src/results'

with open(train_model_path + '/fizz_buzz_NN.pkl', "rb") as model_file:
    fb_network = pickle.load(model_file)

def predict() -> int:
    num_correct = 0
    df = pd.DataFrame(columns=["Número", "Predicción", "Valor real"])
    for i in range(1, 101):
        x = Util().binary_encoder(i)
        predicted = Util().pos_max(fb_network.feed_forward(x)[-1])
        real = Util().pos_max(Util().f_buzz_encoder(i))
        labels = [str(i), "fizz", "buzz", "fizzbuzz"]
        row = pd.DataFrame({'Número': [i], 'Predicción': [labels[predicted]], 'Valor real':[labels[real]]})
        df = pd.concat([df, row])
        if (predicted == real):
            num_correct += 1
    return num_correct, df

def export_results(df):
     df.to_excel(results_path+'/results_predictions.xlsx', index=False)
     wb = load_workbook(results_path+'/results_predictions.xlsx')
     ws = wb.active
     color = PatternFill(start_color="FF0000", fill_type="solid")
     for row in range(2, ws.max_row + 1):
            cell_1 = ws.cell(row=row, column=2)
            cell_2 = ws.cell(row=row, column=3)

            if cell_1.value != cell_2.value:
                cell_1.fill = color
                cell_2.fill = color
     wb.save(results_path+'/results_predictions.xlsx')
     return

if __name__ == "__main__":
    result, df = predict()
    print(f"Resultados correctos: {result}/100")
    Config().create_dir(dir_path=results_path)
    export_results(df)
   

